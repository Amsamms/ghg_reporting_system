"""
Unit Tests for GHGExcelGenerator Module

This module contains comprehensive unit tests for the GHGExcelGenerator class,
ensuring all functionality works correctly and handles edge cases properly.
"""

import pytest
import pandas as pd
import os
import tempfile
from pathlib import Path
import openpyxl
from datetime import datetime

from excel_generator import GHGExcelGenerator


class TestGHGExcelGenerator:
    """Test suite for GHGExcelGenerator class"""

    @pytest.fixture
    def generator(self):
        """Create a GHGExcelGenerator instance for testing"""
        return GHGExcelGenerator()

    @pytest.mark.unit
    def test_initialization(self, generator):
        """Test proper initialization of GHGExcelGenerator"""
        assert generator.company_info is not None
        assert 'name' in generator.company_info
        assert 'reporting_year' in generator.company_info
        assert 'report_date' in generator.company_info
        assert 'facilities' in generator.company_info

        assert generator.company_info['name'] == 'PetrolCorp International'
        assert generator.company_info['reporting_year'] == 2024
        assert len(generator.company_info['facilities']) == 4

    @pytest.mark.unit
    def test_generate_dummy_data_structure(self, generator):
        """Test the structure of generated dummy data"""
        data = generator.generate_dummy_data()

        # Check main data structure
        assert isinstance(data, dict)
        expected_keys = ['scope1', 'scope2', 'scope3', 'energy', 'facilities', 'totals']
        for key in expected_keys:
            assert key in data, f"Missing key: {key}"

        # Check scope1 data
        assert isinstance(data['scope1'], list)
        assert len(data['scope1']) == 9  # 9 scope1 sources
        for item in data['scope1']:
            assert 'Source' in item
            assert 'Annual_Total' in item
            assert 'Percentage' in item
            # Check monthly data
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for month in months:
                assert month in item

        # Check scope2 data
        assert isinstance(data['scope2'], list)
        assert len(data['scope2']) == 3  # 3 scope2 sources

        # Check scope3 data
        assert isinstance(data['scope3'], list)
        assert len(data['scope3']) == 12  # 12 scope3 sources

        # Check energy data
        assert isinstance(data['energy'], list)
        assert len(data['energy']) == 6  # 6 energy sources

        # Check facilities data
        assert isinstance(data['facilities'], list)
        assert len(data['facilities']) == 4  # 4 facilities

        # Check totals
        assert isinstance(data['totals'], dict)
        assert 'scope1_total' in data['totals']
        assert 'scope2_total' in data['totals']
        assert 'scope3_total' in data['totals']
        assert 'grand_total' in data['totals']

    @pytest.mark.unit
    def test_generate_dummy_data_values(self, generator):
        """Test the validity of generated data values"""
        data = generator.generate_dummy_data()

        # Test scope1 data values
        for item in data['scope1']:
            assert isinstance(item['Annual_Total'], (int, float))
            assert item['Annual_Total'] > 0
            assert isinstance(item['Percentage'], (int, float))
            assert 0 <= item['Percentage'] <= 100

        # Test scope2 data values
        for item in data['scope2']:
            assert isinstance(item['Annual_Total'], (int, float))
            assert item['Annual_Total'] > 0

        # Test scope3 data values
        for item in data['scope3']:
            assert isinstance(item['Annual_Total'], (int, float))
            assert item['Annual_Total'] > 0

        # Test energy data values
        for item in data['energy']:
            assert isinstance(item['Annual_Total'], (int, float))
            assert item['Annual_Total'] > 0
            assert isinstance(item['Emission_Factor'], (int, float))
            assert 0.2 <= item['Emission_Factor'] <= 0.8

        # Test facilities data values
        for item in data['facilities']:
            assert isinstance(item['Scope_1'], (int, float))
            assert isinstance(item['Scope_2'], (int, float))
            assert isinstance(item['Scope_3'], (int, float))
            assert isinstance(item['Energy_Intensity'], (int, float))
            assert isinstance(item['Production'], (int, float))
            assert item['Scope_1'] > 0
            assert item['Scope_2'] > 0
            assert item['Scope_3'] > 0
            assert item['Energy_Intensity'] > 0
            assert item['Production'] > 0

        # Test totals calculation
        scope1_sum = sum(item['Annual_Total'] for item in data['scope1'])
        scope2_sum = sum(item['Annual_Total'] for item in data['scope2'])
        scope3_sum = sum(item['Annual_Total'] for item in data['scope3'])

        assert abs(data['totals']['scope1_total'] - scope1_sum) < 0.01
        assert abs(data['totals']['scope2_total'] - scope2_sum) < 0.01
        assert abs(data['totals']['scope3_total'] - scope3_sum) < 0.01

        expected_grand_total = scope1_sum + scope2_sum + scope3_sum
        assert abs(data['totals']['grand_total'] - expected_grand_total) < 0.01

    @pytest.mark.unit
    def test_percentage_calculation(self, generator):
        """Test that percentages are calculated correctly"""
        data = generator.generate_dummy_data()

        # Test scope1 percentages
        scope1_total = data['totals']['scope1_total']
        for item in data['scope1']:
            expected_percentage = (item['Annual_Total'] / scope1_total) * 100
            assert abs(item['Percentage'] - expected_percentage) < 0.01

        # Test scope2 percentages
        scope2_total = data['totals']['scope2_total']
        for item in data['scope2']:
            expected_percentage = (item['Annual_Total'] / scope2_total) * 100
            assert abs(item['Percentage'] - expected_percentage) < 0.01

        # Test scope3 percentages
        scope3_total = data['totals']['scope3_total']
        for item in data['scope3']:
            expected_percentage = (item['Annual_Total'] / scope3_total) * 100
            assert abs(item['Percentage'] - expected_percentage) < 0.01

    @pytest.mark.unit
    def test_create_excel_template_file_creation(self, generator, temp_output_dir):
        """Test that Excel template file is created correctly"""
        output_file = temp_output_dir / 'test_template.xlsx'

        result = generator.create_excel_template(str(output_file))

        assert result == str(output_file)
        assert output_file.exists()
        assert output_file.stat().st_size > 0

    @pytest.mark.unit
    def test_create_excel_template_sheets(self, generator, temp_output_dir):
        """Test that all required Excel sheets are created"""
        output_file = temp_output_dir / 'test_template.xlsx'
        generator.create_excel_template(str(output_file))

        # Load the Excel file and check sheets
        excel_data = pd.read_excel(output_file, sheet_name=None)

        expected_sheets = [
            'Dashboard', 'Scope 1 Emissions', 'Scope 2 Emissions',
            'Scope 3 Emissions', 'Energy Consumption', 'Facility Breakdown',
            'Targets & Performance'
        ]

        for sheet in expected_sheets:
            assert sheet in excel_data, f"Missing sheet: {sheet}"
            assert not excel_data[sheet].empty, f"Empty sheet: {sheet}"

    @pytest.mark.unit
    def test_create_excel_template_data_integrity(self, generator, temp_output_dir):
        """Test data integrity in created Excel template"""
        output_file = temp_output_dir / 'test_template.xlsx'
        generator.create_excel_template(str(output_file))

        excel_data = pd.read_excel(output_file, sheet_name=None)

        # Check Scope 1 Emissions sheet
        scope1_df = excel_data['Scope 1 Emissions']
        assert 'Source' in scope1_df.columns
        assert 'Annual_Total' in scope1_df.columns
        assert 'Percentage' in scope1_df.columns
        assert len(scope1_df) == 9  # 9 scope1 sources

        # Check monthly columns
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        for month in months:
            assert month in scope1_df.columns

        # Check Scope 2 Emissions sheet
        scope2_df = excel_data['Scope 2 Emissions']
        assert 'Source' in scope2_df.columns
        assert 'Annual_Total' in scope2_df.columns
        assert len(scope2_df) == 3  # 3 scope2 sources

        # Check Scope 3 Emissions sheet
        scope3_df = excel_data['Scope 3 Emissions']
        assert 'Source' in scope3_df.columns
        assert 'Annual_Total' in scope3_df.columns
        assert len(scope3_df) == 12  # 12 scope3 sources

        # Check Energy Consumption sheet
        energy_df = excel_data['Energy Consumption']
        assert 'Energy_Source' in energy_df.columns
        assert 'Annual_Total' in energy_df.columns
        assert 'Emission_Factor' in energy_df.columns
        assert len(energy_df) == 6  # 6 energy sources

        # Check Facility Breakdown sheet
        facility_df = excel_data['Facility Breakdown']
        assert 'Facility' in facility_df.columns
        assert 'Scope_1' in facility_df.columns
        assert 'Scope_2' in facility_df.columns
        assert 'Scope_3' in facility_df.columns
        assert 'Energy_Intensity' in facility_df.columns
        assert 'Production' in facility_df.columns
        assert len(facility_df) == 4  # 4 facilities

        # Check Targets & Performance sheet
        targets_df = excel_data['Targets & Performance']
        assert 'Metric' in targets_df.columns
        assert 'Target_2024' in targets_df.columns
        assert 'Actual_2024' in targets_df.columns
        assert 'Target_2025' in targets_df.columns
        assert 'Status' in targets_df.columns
        assert len(targets_df) == 5  # 5 target metrics

    @pytest.mark.unit
    def test_excel_formatting(self, generator, temp_output_dir):
        """Test that Excel file formatting is applied correctly"""
        output_file = temp_output_dir / 'test_template.xlsx'
        generator.create_excel_template(str(output_file))

        # Load workbook to check formatting
        wb = openpyxl.load_workbook(output_file)

        # Check that all sheets exist
        expected_sheets = [
            'Dashboard', 'Scope 1 Emissions', 'Scope 2 Emissions',
            'Scope 3 Emissions', 'Energy Consumption', 'Facility Breakdown',
            'Targets & Performance'
        ]

        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames
            ws = wb[sheet_name]

            # Check that there's content in the sheet
            assert ws.max_row > 0
            assert ws.max_column > 0

            # Check column width adjustment (should be reasonable)
            for column in ws.columns:
                column_letter = column[0].column_letter
                width = ws.column_dimensions[column_letter].width
                if width:  # Width might be None for default
                    assert 0 < width <= 50  # Reasonable width range

    @pytest.mark.unit
    def test_monthly_data_consistency(self, generator):
        """Test that monthly data sums to annual totals"""
        data = generator.generate_dummy_data()

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        # Test scope1 data
        for item in data['scope1']:
            monthly_sum = sum(item[month] for month in months)
            assert abs(monthly_sum - item['Annual_Total']) < 0.01

        # Test scope2 data
        for item in data['scope2']:
            monthly_sum = sum(item[month] for month in months)
            assert abs(monthly_sum - item['Annual_Total']) < 0.01

        # Test scope3 data
        for item in data['scope3']:
            monthly_sum = sum(item[month] for month in months)
            assert abs(monthly_sum - item['Annual_Total']) < 0.01

        # Test energy data
        for item in data['energy']:
            monthly_sum = sum(item[month] for month in months)
            assert abs(monthly_sum - item['Annual_Total']) < 0.01

    @pytest.mark.unit
    def test_facility_data_consistency(self, generator):
        """Test that facility data is consistent"""
        data = generator.generate_dummy_data()

        for facility in data['facilities']:
            # Check that all facilities have required fields
            assert 'Facility' in facility
            assert 'Scope_1' in facility
            assert 'Scope_2' in facility
            assert 'Scope_3' in facility
            assert 'Energy_Intensity' in facility
            assert 'Production' in facility

            # Check facility name is in the configured list
            assert facility['Facility'] in generator.company_info['facilities']

            # Check value ranges are reasonable
            assert 8000 <= facility['Scope_1'] <= 25000
            assert 3000 <= facility['Scope_2'] <= 12000
            assert 5000 <= facility['Scope_3'] <= 18000
            assert 2.5 <= facility['Energy_Intensity'] <= 8.0
            assert 50000 <= facility['Production'] <= 200000

    @pytest.mark.error_handling
    def test_create_excel_template_invalid_path(self, generator):
        """Test error handling for invalid file paths"""
        # Test with invalid directory
        invalid_path = "/invalid/directory/test.xlsx"

        # Should not raise exception but may fail silently or create directory
        try:
            result = generator.create_excel_template(invalid_path)
            # If it succeeds, check the result
            if result == invalid_path:
                assert os.path.exists(invalid_path)
        except (PermissionError, FileNotFoundError, OSError):
            # Expected for invalid paths
            pass

    @pytest.mark.error_handling
    def test_create_excel_template_readonly_path(self, generator, temp_output_dir):
        """Test error handling for read-only paths"""
        output_file = temp_output_dir / 'readonly_test.xlsx'

        # Create and make readonly
        output_file.touch()
        output_file.chmod(0o444)  # Read-only

        try:
            generator.create_excel_template(str(output_file))
        except PermissionError:
            # Expected for read-only files
            pass
        finally:
            # Cleanup - restore write permissions
            try:
                output_file.chmod(0o644)
                output_file.unlink()
            except:
                pass

    @pytest.mark.unit
    def test_data_randomness(self, generator):
        """Test that data generation includes proper randomness"""
        data1 = generator.generate_dummy_data()
        data2 = generator.generate_dummy_data()

        # Data should be different between runs (random seed not fixed in production)
        # Check at least some values are different
        scope1_values1 = [item['Annual_Total'] for item in data1['scope1']]
        scope1_values2 = [item['Annual_Total'] for item in data2['scope1']]

        # Allow for the possibility that some values might be the same by chance
        # but require that not all values are identical
        differences = [abs(v1 - v2) for v1, v2 in zip(scope1_values1, scope1_values2)]
        assert any(diff > 0.01 for diff in differences), "Generated data should have some randomness"

    @pytest.mark.unit
    def test_source_names_validity(self, generator):
        """Test that all source names are valid and expected"""
        data = generator.generate_dummy_data()

        # Check scope1 source names
        expected_scope1_sources = [
            'Combustion - Natural Gas', 'Combustion - Fuel Oil', 'Combustion - Diesel',
            'Process Emissions - Refining', 'Fugitive - Equipment Leaks', 'Fugitive - Venting',
            'Mobile Combustion - Fleet', 'Flaring', 'Process Venting'
        ]
        actual_scope1_sources = [item['Source'] for item in data['scope1']]
        assert set(actual_scope1_sources) == set(expected_scope1_sources)

        # Check scope2 source names
        expected_scope2_sources = [
            'Purchased Electricity', 'Purchased Steam', 'Purchased Heat/Cooling'
        ]
        actual_scope2_sources = [item['Source'] for item in data['scope2']]
        assert set(actual_scope2_sources) == set(expected_scope2_sources)

        # Check energy source names
        expected_energy_sources = [
            'Natural Gas (MWh)', 'Electricity (MWh)', 'Steam (MWh)',
            'Fuel Oil (MWh)', 'Diesel (MWh)', 'Gasoline (MWh)'
        ]
        actual_energy_sources = [item['Energy_Source'] for item in data['energy']]
        assert set(actual_energy_sources) == set(expected_energy_sources)

    @pytest.mark.unit
    def test_date_format_in_company_info(self, generator):
        """Test that date format in company info is correct"""
        company_info = generator.company_info

        # Check that report_date is in correct format
        try:
            datetime.strptime(company_info['report_date'], '%Y-%m-%d')
        except ValueError:
            pytest.fail("Report date is not in correct YYYY-MM-DD format")

    @pytest.mark.performance
    def test_performance_data_generation(self, generator):
        """Test performance of data generation"""
        import time

        start_time = time.time()
        data = generator.generate_dummy_data()
        end_time = time.time()

        generation_time = end_time - start_time

        # Should complete within reasonable time (adjust threshold as needed)
        assert generation_time < 5.0, f"Data generation took {generation_time:.2f}s, expected < 5.0s"

        # Verify data was actually generated
        assert data is not None
        assert len(data) > 0

    @pytest.mark.performance
    def test_performance_excel_creation(self, generator, temp_output_dir):
        """Test performance of Excel file creation"""
        import time

        output_file = temp_output_dir / 'performance_test.xlsx'

        start_time = time.time()
        result = generator.create_excel_template(str(output_file))
        end_time = time.time()

        creation_time = end_time - start_time

        # Should complete within reasonable time
        assert creation_time < 30.0, f"Excel creation took {creation_time:.2f}s, expected < 30.0s"

        # Verify file was created successfully
        assert result == str(output_file)
        assert output_file.exists()

    @pytest.mark.unit
    def test_reproducibility_with_fixed_seed(self, generator):
        """Test that data generation is reproducible with fixed random seed"""
        import random

        # Set fixed seed
        random.seed(123)
        data1 = generator.generate_dummy_data()

        # Reset same seed
        random.seed(123)
        data2 = generator.generate_dummy_data()

        # Data should be identical
        for i, (item1, item2) in enumerate(zip(data1['scope1'], data2['scope1'])):
            assert abs(item1['Annual_Total'] - item2['Annual_Total']) < 0.001, f"Scope1 item {i} not reproducible"

        for i, (item1, item2) in enumerate(zip(data1['scope2'], data2['scope2'])):
            assert abs(item1['Annual_Total'] - item2['Annual_Total']) < 0.001, f"Scope2 item {i} not reproducible"

if __name__ == "__main__":
    pytest.main([__file__, "-v"])