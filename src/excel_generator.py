import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
import numpy as np
from datetime import datetime, timedelta
import random

class GHGExcelGenerator:
    def __init__(self):
        self.company_info = {
            'name': 'PetrolCorp International',
            'reporting_year': 2024,
            'report_date': datetime.now().strftime('%Y-%m-%d'),
            'facilities': ['Refinery A', 'Refinery B', 'Offshore Platform C', 'Distribution Center D']
        }

    def generate_dummy_data(self):
        """Generate comprehensive dummy GHG data for petroleum company"""
        # Scope 1 Emissions (Direct)
        scope1_sources = [
            'Combustion - Natural Gas', 'Combustion - Fuel Oil', 'Combustion - Diesel',
            'Process Emissions - Refining', 'Fugitive - Equipment Leaks', 'Fugitive - Venting',
            'Mobile Combustion - Fleet', 'Flaring', 'Process Venting'
        ]

        # Scope 2 Emissions (Indirect - Energy)
        scope2_sources = [
            'Purchased Electricity', 'Purchased Steam', 'Purchased Heat/Cooling'
        ]

        # Scope 3 Emissions (Other Indirect)
        scope3_sources = [
            'Purchased Goods/Services', 'Capital Goods', 'Fuel/Energy Activities',
            'Transportation - Upstream', 'Waste Generated', 'Business Travel',
            'Employee Commuting', 'Transportation - Downstream', 'Processing of Products',
            'Use of Sold Products', 'End-of-life Products', 'Leased Assets'
        ]

        # Generate monthly data for each scope
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        # Scope 1 Data (tCO2e)
        scope1_data = []
        for source in scope1_sources:
            monthly_values = [random.uniform(800, 2500) for _ in months]
            scope1_data.append({
                'Source': source,
                'Annual_Total': sum(monthly_values),
                'Percentage': 0,  # Will calculate later
                **dict(zip(months, monthly_values))
            })

        # Scope 2 Data (tCO2e)
        scope2_data = []
        for source in scope2_sources:
            monthly_values = [random.uniform(300, 1200) for _ in months]
            scope2_data.append({
                'Source': source,
                'Annual_Total': sum(monthly_values),
                'Percentage': 0,
                **dict(zip(months, monthly_values))
            })

        # Scope 3 Data (tCO2e)
        scope3_data = []
        for source in scope3_sources:
            monthly_values = [random.uniform(100, 800) for _ in months]
            scope3_data.append({
                'Source': source,
                'Annual_Total': sum(monthly_values),
                'Percentage': 0,
                **dict(zip(months, monthly_values))
            })

        # Calculate percentages
        total_scope1 = sum([row['Annual_Total'] for row in scope1_data])
        total_scope2 = sum([row['Annual_Total'] for row in scope2_data])
        total_scope3 = sum([row['Annual_Total'] for row in scope3_data])
        grand_total = total_scope1 + total_scope2 + total_scope3

        for row in scope1_data:
            row['Percentage'] = (row['Annual_Total'] / total_scope1) * 100
        for row in scope2_data:
            row['Percentage'] = (row['Annual_Total'] / total_scope2) * 100
        for row in scope3_data:
            row['Percentage'] = (row['Annual_Total'] / total_scope3) * 100

        # Emission by source data (energy-related emission sources)
        # Different ranges for each source to show meaningful variation
        emission_sources_ranges = {
            'Natural Gas': (1500, 3000),
            'Electricity': (2000, 3500),
            'Steam': (800, 1500),
            'Fuel Oil': (1000, 2000),
            'Diesel': (600, 1200),
            'Gasoline': (400, 900)
        }

        emission_by_source_data = []
        for source, (min_val, max_val) in emission_sources_ranges.items():
            # Monthly emissions in tCO2e with varied ranges per source
            monthly_values = [random.uniform(min_val, max_val) for _ in months]
            emission_by_source_data.append({
                'Source': source,
                'Annual_Total_tCO2e': sum(monthly_values),
                **dict(zip(months, monthly_values))
            })

        # Facility-wise breakdown
        facility_data = []
        for facility in self.company_info['facilities']:
            facility_data.append({
                'Facility': facility,
                'Scope_1': random.uniform(8000, 25000),
                'Scope_2': random.uniform(3000, 12000),
                'Scope_3': random.uniform(5000, 18000),
                'Energy_Intensity': random.uniform(2.5, 8.0),  # tCO2e/MWh
                'Production': random.uniform(50000, 200000)  # barrels/year
            })

        return {
            'scope1': scope1_data,
            'scope2': scope2_data,
            'scope3': scope3_data,
            'emission_by_source': emission_by_source_data,
            'facilities': facility_data,
            'totals': {
                'scope1_total': total_scope1,
                'scope2_total': total_scope2,
                'scope3_total': total_scope3,
                'grand_total': grand_total
            }
        }

    def create_excel_template(self, filename='ghg_report_template.xlsx'):
        """Create comprehensive Excel template with multiple sheets"""
        data = self.generate_dummy_data()

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Dashboard/Summary Sheet
            summary_data = pd.DataFrame([
                ['Company Name', self.company_info['name']],
                ['Reporting Year', self.company_info['reporting_year']],
                ['Report Date', self.company_info['report_date']],
                ['Total GHG Emissions (tCO2e)', f"{data['totals']['grand_total']:.2f}"],
                ['Scope 1 Emissions (tCO2e)', f"{data['totals']['scope1_total']:.2f}"],
                ['Scope 2 Emissions (tCO2e)', f"{data['totals']['scope2_total']:.2f}"],
                ['Scope 3 Emissions (tCO2e)', f"{data['totals']['scope3_total']:.2f}"],
                ['Total Facilities', len(self.company_info['facilities'])],
                ['Carbon Intensity (tCO2e/barrel)', f"{data['totals']['grand_total']/sum([f['Production'] for f in data['facilities']]):.4f}"]
            ])
            summary_data.to_excel(writer, sheet_name='Dashboard', index=False, header=False)

            # Scope 1 Emissions
            df_scope1 = pd.DataFrame(data['scope1'])
            df_scope1.to_excel(writer, sheet_name='Scope 1 Emissions', index=False)

            # Scope 2 Emissions
            df_scope2 = pd.DataFrame(data['scope2'])
            df_scope2.to_excel(writer, sheet_name='Scope 2 Emissions', index=False)

            # Scope 3 Emissions
            df_scope3 = pd.DataFrame(data['scope3'])
            df_scope3.to_excel(writer, sheet_name='Scope 3 Emissions', index=False)

            # Emission By Source
            df_emission_by_source = pd.DataFrame(data['emission_by_source'])
            df_emission_by_source.to_excel(writer, sheet_name='Emission By Source', index=False)

            # Facility Breakdown
            df_facilities = pd.DataFrame(data['facilities'])
            df_facilities.to_excel(writer, sheet_name='Facility Breakdown', index=False)

            # Targets and Performance
            targets_data = pd.DataFrame([
                {'Metric': 'Total GHG Reduction Target (%)', 'Target_2024': 5, 'Actual_2024': 3.2, 'Target_2025': 10, 'Status': 'On Track'},
                {'Metric': 'Scope 1 Reduction (%)', 'Target_2024': 3, 'Actual_2024': 2.1, 'Target_2025': 7, 'Status': 'Needs Improvement'},
                {'Metric': 'Energy Intensity Reduction (%)', 'Target_2024': 4, 'Actual_2024': 4.5, 'Target_2025': 8, 'Status': 'Exceeded'},
                {'Metric': 'Renewable Energy Usage (%)', 'Target_2024': 15, 'Actual_2024': 12, 'Target_2025': 25, 'Status': 'On Track'},
                {'Metric': 'Carbon Capture Implementation', 'Target_2024': 2, 'Actual_2024': 1, 'Target_2025': 4, 'Status': 'Delayed'}
            ])
            targets_data.to_excel(writer, sheet_name='Targets & Performance', index=False)

        # Format the Excel file
        self._format_excel_file(filename)
        return filename

    def _format_excel_file(self, filename):
        """Apply formatting to the Excel file"""
        wb = openpyxl.load_workbook(filename)

        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Format headers
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

        wb.save(filename)

if __name__ == "__main__":
    generator = GHGExcelGenerator()
    template_path = "/home/amsamms/projects/Amr Abu Mady/ghg_reporting_system/data/ghg_report_template.xlsx"
    generator.create_excel_template(template_path)
    print(f"Excel template created: {template_path}")